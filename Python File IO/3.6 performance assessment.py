from os import chdir
import csv
import openpyxl

chdir(r'C:\Users\ddimo\Local Documents\Python\Local College')

myCSV = open('example.csv', 'r')
reader = csv.reader(myCSV)
fileContents = list(reader)


wb = openpyxl.Workbook()
ws = wb.active

letters = []
for charCode in range(65,91):
    letters.append(chr(charCode))

row = 1
for line in fileContents:
    col = 0
    for cellValue in line:
        ws[letters[col]+str(row)].value = cellValue
        col += 1
    row += 1

for row in range(2, len(fileContents)+1):
    ws[letters[1]+str(row)].value=int(ws[letters[1]+str(row)].value)
    
wb.save('example.xlsx')

wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active
rows = ws.max_row
cols = ws.max_column


lables = openpyxl.chart.Reference(ws,min_col=1, min_row=2, max_row=rows)
data = openpyxl.chart.Reference(ws,min_col=2, min_row=2, max_row=rows, max_col=2)

myChart = openpyxl.chart.BarChart()
myChart.add_data(data)
myChart.set_categories(lables)
myChart.title = 'Fruit Quantities'
myChart.legend = None
myChart.x_axis.title="Fruit Types"
myChart.y_axis.title="Quantities"


ws.add_chart(myChart, 'D1')
wb.save('example.xlsx')

